import  openpyxl

wb = openpyxl.Workbook()
wb.save("files/new_wb.xlsx")

sheet2 = wb.create_sheet(title="Sheet2")

wb.create_sheet(title = "Sheet3")

wb.remove(sheet2)

wb = wb["Sheet3"]
wb.copy_worksheet(ws)

wb.save("files/new_wb2.xlsx")

wb.remove(sheet2)

wb.save("files/new_wb2.xlsx")


